import os
import uuid
import random
import datetime
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key_here'  # Замените на надёжное значение
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mytestingapp.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# МОДЕЛИ

class Faculty(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    subjects = db.relationship('Subject', backref='faculty', lazy=True)

class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.id'), nullable=False)
    questions = db.relationship('Question', backref='subject', lazy=True)

class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    question_text = db.Column(db.Text, nullable=False)
    correct_answer = db.Column(db.String(200), nullable=False)
    answer2 = db.Column(db.String(200), nullable=False)
    answer3 = db.Column(db.String(200), nullable=False)
    answer4 = db.Column(db.String(200), nullable=False)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'admin', 'manager', 'user'
    active = db.Column(db.Boolean, default=True)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.id'), nullable=True)
    session_token = db.Column(db.String(100), nullable=True)
    last_activity = db.Column(db.DateTime, nullable=True)

    faculty = db.relationship('Faculty')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class TestResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    score = db.Column(db.Float, nullable=False)
    correct_count = db.Column(db.Integer, nullable=False)
    total_time = db.Column(db.Integer, nullable=False)  # в секундах

    user = db.relationship('User')
    subject = db.relationship('Subject')

class UserQuestionHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.datetime.utcnow)

    user = db.relationship('User')
    question = db.relationship('Question')


# ДЕКОРАТОРЫ

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("Пожалуйста, войдите в систему", "warning")
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.active:
            flash("Доступ запрещён", "danger")
            return redirect(url_for('login'))
        if 'session_token' not in session or session['session_token'] != user.session_token:
            flash("Ваша сессия завершена, так как вы вошли с другого устройства", "warning")
            return redirect(url_for('logout'))
        user.last_activity = datetime.datetime.utcnow()
        db.session.commit()
        return f(*args, **kwargs)
    return decorated_function

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = User.query.get(session.get('user_id'))
            if user.role != role:
                flash("Доступ запрещён", "danger")
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def roles_required(roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = User.query.get(session.get('user_id'))
            if user.role not in roles:
                flash("Доступ запрещён", "danger")
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Контекстный процессор для текущего пользователя
@app.context_processor
def inject_user():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
    else:
        user = None
    return dict(current_user=user)


# РОУТЫ

@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email').strip()
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password) and user.active:
            token = str(uuid.uuid4())
            user.session_token = token
            user.last_activity = datetime.datetime.utcnow()
            db.session.commit()
            session['user_id'] = user.id
            session['session_token'] = token
            flash("Успешный вход", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("Неверный email или пароль", "danger")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    user = User.query.get(session.get('user_id'))
    if user:
        user.session_token = None
        db.session.commit()
    session.clear()
    flash("Вы вышли из системы", "success")
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    user = User.query.get(session['user_id'])
    if user.role in ['admin', 'manager']:
        return render_template('dashboard.html', user=user)
    else:
        subjects = Subject.query.filter_by(faculty_id=user.faculty_id).all()
        stats = {}
        for subject in subjects:
            results = TestResult.query.filter_by(user_id=user.id, subject_id=subject.id).all()
            if results:
                avg_score = sum(r.score for r in results) / len(results)
                stats[subject] = {'attempts': len(results), 'avg_score': round(avg_score, 2)}
            else:
                stats[subject] = {'attempts': 0, 'avg_score': 0}
        return render_template('dashboard.html', user=user, stats=stats)

# Админ: управление факультетами

@app.route('/admin/faculties', methods=['GET','POST'])
@login_required
@role_required('admin')
def manage_faculties():
    if request.method == 'POST':
        name = request.form.get('name').strip()
        if name:
            faculty = Faculty(name=name)
            db.session.add(faculty)
            db.session.commit()
            flash("Факультет добавлен", "success")
        return redirect(url_for('manage_faculties'))
    faculties = Faculty.query.all()
    return render_template('admin_faculties.html', faculties=faculties)

@app.route('/admin/faculties/delete/<int:faculty_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_faculty(faculty_id):
    faculty = Faculty.query.get_or_404(faculty_id)
    db.session.delete(faculty)
    db.session.commit()
    flash("Факультет удален", "success")
    return redirect(url_for('manage_faculties'))

# Админ: управление предметами

@app.route('/admin/subjects', methods=['GET','POST'])
@login_required
@role_required('admin')
def manage_subjects():
    faculties = Faculty.query.all()
    if request.method == 'POST':
        name = request.form.get('name').strip()
        faculty_id = request.form.get('faculty_id')
        if name and faculty_id:
            subject = Subject(name=name, faculty_id=int(faculty_id))
            db.session.add(subject)
            db.session.commit()
            flash("Предмет добавлен", "success")
        return redirect(url_for('manage_subjects'))
    subjects = Subject.query.all()
    return render_template('admin_subjects.html', subjects=subjects, faculties=faculties)

@app.route('/admin/subjects/delete/<int:subject_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_subject(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    db.session.delete(subject)
    db.session.commit()
    flash("Предмет удален", "success")
    return redirect(url_for('manage_subjects'))

# Админ: загрузка вопросов из Excel

@app.route('/admin/upload_questions', methods=['GET','POST'])
@login_required
@role_required('admin')
def upload_questions():
    if request.method == 'POST':
        subject_id = request.form.get('subject_id')
        file = request.files.get('file')
        if subject_id and file:
            subject = Subject.query.get(int(subject_id))
            if subject:
                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                # Пропускаем заголовок (первая строка)
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    question_text = str(row[0]).strip()
                    correct_answer = str(row[1]).strip()
                    answer2 = str(row[2]).strip()
                    answer3 = str(row[3]).strip()
                    answer4 = str(row[4]).strip()
                    question = Question(
                        subject_id = subject.id,
                        question_text = question_text,
                        correct_answer = correct_answer,
                        answer2 = answer2,
                        answer3 = answer3,
                        answer4 = answer4
                    )
                    db.session.add(question)
                db.session.commit()
                flash("Вопросы успешно загружены", "success")
            else:
                flash("Неверный предмет", "danger")
        else:
            flash("Выберите предмет и файл", "danger")
        return redirect(url_for('upload_questions'))
    subjects = Subject.query.all()
    return render_template('admin_upload_questions.html', subjects=subjects)

# Админ/Менеджер: управление пользователями

@app.route('/admin/users', methods=['GET','POST'])
@login_required
@roles_required(['admin', 'manager'])
def manage_users():
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        name = request.form.get('name').strip()
        email = request.form.get('email').strip()
        password = request.form.get('password')
        role = request.form.get('role')
        faculty_id = request.form.get('faculty_id')
        if user_id:  # обновление пользователя
            user = User.query.get(int(user_id))
            if user:
                user.name = name
                user.email = email
                if password:
                    user.set_password(password)
                user.role = role
                user.faculty_id = int(faculty_id) if faculty_id else None
                db.session.commit()
                flash("Пользователь обновлен", "success")
        else:  # создание нового пользователя
            if name and email and password and role:
                user = User(name=name, email=email, role=role)
                user.set_password(password)
                user.faculty_id = int(faculty_id) if faculty_id else None
                db.session.add(user)
                db.session.commit()
                flash("Пользователь создан", "success")
        return redirect(url_for('manage_users'))
    users = User.query.all()
    faculties = Faculty.query.all()
    return render_template('admin_users.html', users=users, faculties=faculties)

@app.route('/admin/users/deactivate/<int:user_id>', methods=['POST'])
@login_required
@roles_required(['admin', 'manager'])
def deactivate_user(user_id):
    user = User.query.get_or_404(user_id)
    user.active = False
    db.session.commit()
    flash("Пользователь деактивирован", "success")
    return redirect(url_for('manage_users'))

@app.route('/admin/users/activate/<int:user_id>', methods=['POST'])
@login_required
@roles_required(['admin', 'manager'])
def activate_user(user_id):
    user = User.query.get_or_404(user_id)
    user.active = True
    db.session.commit()
    flash("Пользователь активирован", "success")
    return redirect(url_for('manage_users'))

@app.route('/admin/users/terminate/<int:user_id>', methods=['POST'])
@login_required
@roles_required(['admin', 'manager'])
def terminate_user_session(user_id):
    user = User.query.get_or_404(user_id)
    user.session_token = None
    db.session.commit()
    flash("Сессия пользователя завершена", "success")
    return redirect(url_for('manage_users'))

# Тестирование

@app.route('/test/select')
@login_required
def test_select():
    user = User.query.get(session['user_id'])
    subjects = Subject.query.filter_by(faculty_id=user.faculty_id).all()
    return render_template('test_select.html', subjects=subjects)

@app.route('/test/instructions/<mode>/<int:subject_id>')
@login_required
def test_instructions(mode, subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if mode == 'training':
        instructions = "В режиме тренировки будет 25 вопросов, 25 минут. Если вы покинете страницу – тест засчитан не будет."
    else:
        instructions = "В режиме 'Все вопросы' вы увидите ответ сразу после выбора. Изменить ответ нельзя."
    return render_template('test_instructions.html', mode=mode, subject=subject, instructions=instructions)

@app.route('/test/training/<int:subject_id>/start')
@login_required
def start_training_test(subject_id):
    user = User.query.get(session['user_id'])
    subject = Subject.query.get_or_404(subject_id)
    session['test'] = {
        'mode': 'training',
        'subject_id': subject_id,
        'start_time': datetime.datetime.utcnow().isoformat(),
        'questions': [],
        'answers': {}
    }
    questions = Question.query.filter_by(subject_id=subject_id).all()
    if not questions:
        flash("Нет вопросов для выбранного предмета", "danger")
        return redirect(url_for('test_select'))
    question_ids = [q.id for q in questions]
    history_counts = { qid: UserQuestionHistory.query.filter_by(user_id=user.id, question_id=qid).count() for qid in question_ids }
    sorted_questions = sorted(question_ids, key=lambda qid: history_counts.get(qid, 0))
    selected = []
    if len(sorted_questions) >= 25:
        selected = sorted_questions[:25]
    else:
        selected = sorted_questions.copy()
        while len(selected) < 25:
            selected.append(random.choice(sorted_questions))
    random.shuffle(selected)
    session['test']['questions'] = selected
    session['test']['current_index'] = 0
    return redirect(url_for('training_question', subject_id=subject_id, index=0))

@app.route('/test/training/<int:subject_id>/question/<int:index>', methods=['GET','POST'])
@login_required
def training_question(subject_id, index):
    test_session = session.get('test')
    if not test_session or test_session.get('mode') != 'training' or int(test_session.get('subject_id')) != subject_id:
        flash("Тест не инициализирован", "danger")
        return redirect(url_for('test_select'))
    questions_ids = test_session.get('questions')
    if index < 0 or index >= len(questions_ids):
        flash("Неверный номер вопроса", "danger")
        return redirect(url_for('test_select'))
    question_id = questions_ids[index]
    question = Question.query.get(question_id)
    answers = [question.correct_answer, question.answer2, question.answer3, question.answer4]
    random.shuffle(answers)
    if request.method == 'POST':
        selected_answer = request.form.get('answer')
        test_session['answers'][str(question_id)] = selected_answer
        session['test'] = test_session
        if 'next' in request.form:
            next_index = index + 1
            if next_index >= len(questions_ids):
                return redirect(url_for('finish_training_test', subject_id=subject_id))
            else:
                return redirect(url_for('training_question', subject_id=subject_id, index=next_index))
        elif 'prev' in request.form:
            prev_index = index - 1
            if prev_index < 0:
                prev_index = 0
            return redirect(url_for('training_question', subject_id=subject_id, index=prev_index))
        elif 'finish' in request.form:
            return redirect(url_for('finish_training_test', subject_id=subject_id))
    start_time = datetime.datetime.fromisoformat(test_session.get('start_time'))
    elapsed = (datetime.datetime.utcnow() - start_time).total_seconds()
    remaining = max(25*60 - int(elapsed), 0)
    if remaining <= 0:
        return redirect(url_for('finish_training_test', subject_id=subject_id))
    return render_template('test_question.html', question=question, index=index, total=len(questions_ids),
                           selected=test_session['answers'].get(str(question_id)), remaining=remaining,
                           answers=answers, mode='training')

@app.route('/test/training/<int:subject_id>/finish')
@login_required
def finish_training_test(subject_id):
    test_session = session.get('test')
    if not test_session or test_session.get('mode') != 'training':
        flash("Тест не инициализирован", "danger")
        return redirect(url_for('test_select'))
    user = User.query.get(session['user_id'])
    questions_ids = test_session.get('questions')
    answers = test_session.get('answers')
    correct = 0
    for qid in questions_ids:
        question = Question.query.get(qid)
        chosen = answers.get(str(qid))
        if chosen == question.correct_answer:
            correct += 1
        history = UserQuestionHistory(user_id=user.id, question_id=qid)
        db.session.add(history)
    total_questions = len(questions_ids)
    score = (correct / total_questions) * 100
    start_time = datetime.datetime.fromisoformat(test_session.get('start_time'))
    total_time = int((datetime.datetime.utcnow() - start_time).total_seconds())
    result = TestResult(user_id=user.id, subject_id=subject_id, score=score, correct_count=correct, total_time=total_time)
    db.session.add(result)
    db.session.commit()
    session.pop('test', None)
    return render_template('test_result.html', correct=correct, total=total_questions, score=round(score,2), total_time=total_time)

@app.route('/test/all/<int:subject_id>/start')
@login_required
def start_all_questions_test(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    questions = Question.query.filter_by(subject_id=subject_id).all()
    questions_ids = [q.id for q in questions]
    random.shuffle(questions_ids)
    session['test'] = {
        'mode': 'all',
        'subject_id': subject_id,
        'questions': questions_ids,
        'answers': {}
    }
    session['test']['current_index'] = 0
    return redirect(url_for('all_question', subject_id=subject_id, index=0))

@app.route('/test/all/<int:subject_id>/question/<int:index>', methods=['GET','POST'])
@login_required
def all_question(subject_id, index):
    test_session = session.get('test')
    if not test_session or test_session.get('mode') != 'all' or int(test_session.get('subject_id')) != subject_id:
        flash("Тест не инициализирован", "danger")
        return redirect(url_for('test_select'))
    questions_ids = test_session.get('questions')
    if index < 0 or index >= len(questions_ids):
        flash("Неверный номер вопроса", "danger")
        return redirect(url_for('test_select'))
    question_id = questions_ids[index]
    question = Question.query.get(question_id)
    answers = [question.correct_answer, question.answer2, question.answer3, question.answer4]
    random.shuffle(answers)
    feedback = None
    if request.method == 'POST':
        selected_answer = request.form.get('answer')
        if selected_answer == question.correct_answer:
            feedback = "Правильно"
        else:
            feedback = f"Неправильно. Правильный ответ: {question.correct_answer}"
        test_session['answers'][str(question_id)] = selected_answer
        session['test'] = test_session
    return render_template('all_questions.html', question=question, index=index, total=len(questions_ids),
                           answers=answers, feedback=feedback)

@app.route('/support')
def support():
    return render_template('support.html')

# Инициализация БД и создание дефолтного администратора
@app.before_first_request
def create_tables():
    db.create_all()
    admin = User.query.filter_by(role='admin').first()
    if not admin:
        admin = User(name="Администратор", email="admin@example.com", role="admin")
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()

if __name__ == '__main__':
    app.run(debug=True)
